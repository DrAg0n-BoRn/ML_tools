import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from typing import List, Optional


def split_sheets_excel(file_path: str):
    """
    Splits a multi-sheet Excel file into separate Excel files per sheet which are saved in a subdirectory.

    Parameters:
        file_path (str): Path to the target Excel file.
        
    Returns:
        output_dir (str): Path to output directory.
    """
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    
    base_dir = os.path.dirname(os.path.abspath(file_path))
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_dir = os.path.join(base_dir, f"{base_name}_split_sheets")
    os.makedirs(output_dir, exist_ok=True)

    count = 0
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        for row in ws.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

        output_filename = f"{base_name}_{sheet_name}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        new_wb.save(output_path)
        count += 1

    print(f"Created {count} Excel files (one for each sheet).")
    return output_dir


def split_sheets_from_directory(input_dir: str, output_dir: str) -> None:
    """
    Processes all Excel files in `input_dir` by splitting each sheet into a separate Excel file
    and saving them in `output_dir`.

    Parameters:
        input_dir (str): Path to the directory containing input Excel files.
        output_dir (str): Path to the directory where split files will be saved.
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    input_files = [
        f for f in os.listdir(input_dir)
        if f.endswith('.xlsx') and not f.startswith('~')
    ]
    
    if not input_files:
        raise FileNotFoundError("No Excel files found in the input directory.")

    output_file_count = 0

    for filename in input_files:
        file_path = os.path.join(input_dir, filename)
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        base_name = os.path.splitext(filename)[0]

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            for row in ws.iter_rows():
                for cell in row:
                    new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

            output_filename = f"{base_name}_{sheet_name}.xlsx"
            output_path = os.path.join(output_dir, output_filename)
            new_wb.save(output_path)
            output_file_count += 1

    print(f"Processed {len(input_files)} input Excel file(s).")
    print(f"Generated {output_file_count} output Excel file(s).")
    return None


def unmerge_columns_excel(file_path: str) -> None:
    """
    Unmerges vertically merged cells of an Excel file. Must contain exactly one worksheet.
    Fills down the top-left value for each vertically merged column range.
    Saves the modified file in the same directory with '_unmerged' suffix.

    Parameters:
        file_path (str): Path to the target Excel file.

    Returns:
        None
    """
    def _process_single_sheet(src_ws: Worksheet) -> Worksheet:
        new_ws = Workbook().active
        new_ws.title = src_ws.title

        for row in src_ws.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

        for merged_range in list(src_ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = (
                merged_range.min_row, merged_range.min_col,
                merged_range.max_row, merged_range.max_col
            )
            if min_col == max_col:  # Vertical merge in a single column
                value = src_ws.cell(row=min_row, column=min_col).value
                for row in range(min_row, max_row + 1):
                    new_ws.cell(row=row, column=min_col, value=value)
            new_ws.unmerge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col
            )

        return new_ws

    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    if len(sheet_names) != 1:
        raise ValueError("The workbook must contain exactly one worksheet.")

    src_ws = wb[sheet_names[0]]
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    processed_ws = _process_single_sheet(src_ws)
    new_wb._add_sheet(processed_ws)

    base_dir = os.path.dirname(os.path.abspath(file_path))
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_path = os.path.join(base_dir, f"{base_name}_unmerged.xlsx")
    new_wb.save(output_path)

    print(f"Unmerged and saved: {output_path}")
    return None


def unmerge_columns_from_directory(input_dir: str, output_dir: str) -> None:
    """
    Processes all Excel files in `input_dir` by unmerging vertically merged columns. Each file must contain exactly one worksheet.
    Saves modified files with '_unmerged' suffix in output_dir.

    Parameters:
        input_dir (str): Path to the directory containing input Excel files.
        output_dir (str): Path to the directory for output Excel files.

    Returns:
        None
    """
    def _process_single_sheet(src_ws: Worksheet) -> Worksheet:
        new_ws = Workbook().active
        new_ws.title = src_ws.title

        for row in src_ws.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

        for merged_range in list(src_ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = (
                merged_range.min_row, merged_range.min_col,
                merged_range.max_row, merged_range.max_col
            )
            if min_col == max_col:
                value = src_ws.cell(row=min_row, column=min_col).value
                for row in range(min_row, max_row + 1):
                    new_ws.cell(row=row, column=min_col, value=value)
            new_ws.unmerge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col
            )

        return new_ws

    os.makedirs(output_dir, exist_ok=True)

    input_files = [
        f for f in os.listdir(input_dir)
        if f.endswith('.xlsx') and not f.startswith('~')
    ]

    if not input_files:
        raise FileNotFoundError("No Excel files found in the input directory.")

    output_count = 0

    for filename in input_files:
        file_path = os.path.join(input_dir, filename)
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        if len(sheet_names) != 1:
            print(f"Skipping '{filename}' (expected exactly one sheet).")
            continue

        src_ws = wb[sheet_names[0]]
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        processed_ws = _process_single_sheet(src_ws)
        new_wb._add_sheet(processed_ws)

        base_name = os.path.splitext(filename)[0]
        output_path = os.path.join(output_dir, f"{base_name}_unmerged.xlsx")
        new_wb.save(output_path)
        output_count += 1
        print(f"Processed: {filename}")

    print(f"Processed {len(input_files)} input file(s).")
    print(f"Generated {output_count} output file(s).")
    return None


def validate_excel_schema(
    target_dir: str,
    expected_columns: List[str],
    strict: bool = False
) -> List[str]:
    """
    Validates that each Excel file in a directory conforms to the expected column schema.
    
    Parameters:
        target_dir (str): Path to the directory containing Excel files.
        expected_columns (list[str]): List of expected column names.
        strict (bool): If True, columns must match exactly (names and order).
                      If False, columns must contain at least all expected names.

    Returns:
        List[str]: List of file paths that failed the schema validation.
    """
    invalid_files = []
    expected_set = set(expected_columns)
    
    excel_seen = 0

    for filename in os.listdir(target_dir):
        if not filename.lower().endswith(".xlsx"):
            continue  # Skip non-Excel files
        
        if filename.startswith("~"): # Skip temporary files
            continue

        file_path = os.path.join(target_dir, filename)
        excel_seen += 1
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active  # Only check the first worksheet

            header = [cell.value for cell in next(ws.iter_rows(max_row=1))]

            if strict:
                if header != expected_columns:
                    invalid_files.append(file_path)
            else:
                header_set = set(header)
                if not expected_set.issubset(header_set):
                    invalid_files.append(file_path)

        except Exception as e:
            print(f"Error processing '{file_path}': {e}")
            invalid_files.append(file_path)
    
    valid_excel_number = excel_seen - len(invalid_files)
    print(f"{valid_excel_number} excel files conform to the schema.")
    if invalid_files:
        print(f"{len(invalid_files)} excel files are invalid.")

    return invalid_files


def vertical_merge_transform_excel(
    target_dir: str,
    csv_filename: str,
    output_dir: str,
    target_columns: Optional[List[str]] = None,
    rename_columns: Optional[List[str]] = None
) -> None:
    """
    Merges multiple Excel files in a directory vertically and saves as a single CSV file.

    Constraints:
    - Only the first worksheet of each Excel file is processed.
    - All Excel files must have either the same column names or a common subset if `target_columns` is provided.
    - If `rename_columns` is provided, it must match the length of `target_columns` (if used) or the original columns. Names match by position.

    Parameters:
        target_dir (str): Directory containing Excel files.
        csv_filename (str): Output CSV filename.
        output_dir (str): Directory to save the output CSV file.
        target_columns (list[str] | None): Columns to select from each Excel file.
        rename_columns (list[str] | None): Optional renaming for columns. Position-based matching.
        
    Returns:
        csv_path (str): 
    """
    raw_excel_files = [f for f in os.listdir(target_dir) if f.endswith(('.xlsx', '.xls'))]
    excel_files = [f for f in raw_excel_files if not f.startswith('~')]  # Exclude temporary files
    
    if not excel_files:
        raise ValueError("No Excel files found in the target directory.")

    csv_filename = csv_filename if csv_filename.endswith('.csv') else f"{csv_filename}.csv"
    csv_path = os.path.join(output_dir, csv_filename)

    dataframes = []
    for file in excel_files:
        file_path = os.path.join(target_dir, file)
        df = pd.read_excel(file_path, engine='openpyxl')

        if target_columns is not None:
            missing = [col for col in target_columns if col not in df.columns]
            if missing:
                raise ValueError(f"Missing columns in {file}: {missing}")
            df = df[target_columns]

        dataframes.append(df)

    merged_df = pd.concat(dataframes, ignore_index=True)

    if rename_columns is not None:
        expected_len = len(target_columns if target_columns is not None else merged_df.columns)
        if len(rename_columns) != expected_len:
            raise ValueError("Length of rename_columns must match the selected columns")
        merged_df.columns = rename_columns

    merged_df.to_csv(csv_path, index=False, encoding='utf-8')
    print(f"✅ Merged {len(dataframes)} excel files into '{csv_filename}'.")


def horizontal_merge_transform_excel(
    target_dir: str,
    csv_filename: str,
    output_dir: str,
    drop_columns: Optional[list[str]] = None,
    skip_duplicates: bool = False
) -> None:
    """
    Horizontally concatenates Excel files (first sheet of each) by aligning rows and expanding columns. 
    Then saves the result as a .csv file.

    Constraints:
    - All Excel files must have the same number of rows, or shorter ones will be padded with empty rows.
    - Only the first sheet in each Excel file is used.
    - Columns in `drop_columns` will be excluded from the result.
    - If `skip_duplicates` is False, duplicate columns are suffixed with "_copy", "_copy2", etc.
      If True, only the first occurrence of each column name is kept.

    Parameters:
        target_dir (str): Directory containing Excel files.
        csv_filename (str): Name of the output CSV file.
        output_dir (str): Directory to save the output CSV file.
        drop_columns (list[str] | None): Columns to exclude from each file before merging.
        skip_duplicates (bool): Whether to skip duplicate columns or rename them.
    """
    raw_excel_files = [f for f in os.listdir(target_dir) if f.endswith(('.xlsx', '.xls'))]
    excel_files = [f for f in raw_excel_files if not f.startswith('~')]  # Exclude temporary files
    if not excel_files:
        raise ValueError("No Excel files found in the target directory.")

    csv_filename = csv_filename if csv_filename.endswith('.csv') else f"{csv_filename}.csv"
    csv_path = os.path.join(output_dir, csv_filename)

    dataframes = []
    max_rows = 0

    for file in excel_files:
        file_path = os.path.join(target_dir, file)
        df = pd.read_excel(file_path, engine='openpyxl')

        if drop_columns is not None:
            df = df.drop(columns=[col for col in drop_columns if col in df.columns])

        max_rows = max(max_rows, len(df))
        dataframes.append(df)

    padded_dataframes = []
    for df in dataframes:
        padded_df = df.reindex(range(max_rows)).reset_index(drop=True)
        padded_dataframes.append(padded_df)

    merged_df = pd.concat(padded_dataframes, axis=1)

    duplicate_columns = merged_df.columns[merged_df.columns.duplicated()].tolist()

    if skip_duplicates:
        merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
    else:
        seen = {}
        new_cols = []
        for col in merged_df.columns:
            base_col = col
            count = seen.get(base_col, 0)
            if count:
                while f"{base_col}_copy{count}" in seen:
                    count += 1
                col = f"{base_col}_copy{count}"
            seen[col] = count + 1
            new_cols.append(col)
        merged_df.columns = new_cols

    merged_df.to_csv(csv_path, index=False, encoding='utf-8')

    print(f"✅ Merged {len(excel_files)} Excel files into '{csv_filename}'.")
    if duplicate_columns:
        print(f"⚠️ Duplicate columns: {duplicate_columns}")

